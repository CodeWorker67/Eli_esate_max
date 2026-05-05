import asyncio
import html as html_lib

from maxapi.context.base import BaseContext
from maxapi.context.state_machine import State, StatesGroup
from maxapi.enums.attachment import AttachmentType
from maxapi.enums.parse_mode import ParseMode
from maxapi.filters.command import Command, CommandStart
from maxapi.types.attachments.buttons.callback_button import CallbackButton
from maxapi.types.attachments.file import File
from maxapi.types.updates.message_callback import MessageCallback
from maxapi.types.updates.message_created import MessageCreated
from maxapi.utils.inline_keyboard import InlineKeyboardBuilder
import aiohttp
import openpyxl
from io import BytesIO
from maxapi import F, Router
from sqlalchemy import delete, func, select

from bot import bot
from max_helpers import (
    answer_message,
    inline_kb,
    callback_ack,
    edit_or_send_callback,
    fio_html,
    profile_link_line_html,
    send_user,
    states_in_group,
    text_from_message,
    user_id_from_message,
)
from config import ADMIN_IDS, PAGE_SIZE, RAZRAB
from db.models import (
    Appeal,
    AsyncSessionLocal,
    Contractor,
    ContractorContractorRequest,
    ContractorRegistrationRequest,
    Manager,
    PermanentPass,
    RegistrationRequest,
    Resident,
    ResidentContractorRequest,
    Security,
    TempPassYooKassaPayment,
    TemporaryPass,
)
from filters import IsAdmin, IsAdminOrManager

router = Router(router_id="admin_user_management")
router.filter(IsAdminOrManager())


def _resident_list_button_text(resident: Resident) -> str:
    base = (resident.fio or "").strip() or f"ID {resident.id}"
    if len(base) > 30:
        return base[:27] + "..."
    return base


def _contractor_list_button_text(contractor: Contractor) -> str:
    company = (contractor.company or "").strip()
    position = (contractor.position or "").strip()
    base = "_".join(p for p in (company, position) if p) or (contractor.fio or "").strip() or f"ID {contractor.id}"
    if len(base) > 30:
        return base[:27] + "..."
    return base


def _parse_list_page(payload: str, prefix: str) -> int | None:
    """prefix: 'list_residents' или 'list_contractors'."""
    if payload == prefix:
        return 0
    suffix = f"{prefix}_p_"
    if payload.startswith(suffix):
        try:
            return int(payload.split("_")[-1])
        except ValueError:
            return 0
    return None


class AddUserStates(StatesGroup):
    WAITING_PHONE = State()
    CHOOSE_TYPE = State()


class ExportStates(StatesGroup):
    WAITING_FILE = State()


@router.message_callback(F.callback.payload == "back_to_main")
async def back_to_main_menu(event: MessageCallback):
    uid = event.callback.user.user_id
    kb = get_admin_menu() if uid in ADMIN_IDS else get_manager_menu()
    await edit_or_send_callback(bot, event, "Добро пожаловать в Главное меню", kb)



def is_valid_phone(phone: str) -> bool:
    return len(phone) == 11 and phone.isdigit() and phone[0] == '8'


# Обновленное главное меню админа
def get_admin_menu():
    return inline_kb([
        [CallbackButton(text="👥Управление пользователями", payload="user_management")],
        [CallbackButton(text="📝 Регистрация", payload="registration_menu")],
        [CallbackButton(text="🚪 Пропуска", payload="passes_menu")],
        [CallbackButton(text="🔍 Поиск пропуска", payload="search_pass")],
        [CallbackButton(text="📈Статистика", payload="statistics_menu")],
        [CallbackButton(text="📩 Выполнить рассылку", payload="posting")],
    ])


def get_manager_menu():
    """Главное меню менеджера: те же разделы, что у админа; в «Управление пользователями» — только резиденты и подрядчики."""
    return get_admin_menu()


def get_admin_user_management_menu():
    return inline_kb([
        [CallbackButton(text="Менеджеры", payload="managers_manage")],
        [CallbackButton(text="СБ", payload="security_manage")],
        [CallbackButton(text="Резиденты", payload="residents_manage")],
        [CallbackButton(text="Подрядчики", payload="contractors_manage")],
        [CallbackButton(text="⬅️ Назад", payload="back_to_main")]
        ])


def get_manager_user_management_menu():
    return inline_kb([
        [CallbackButton(text="Резиденты", payload="residents_manage")],
        [CallbackButton(text="Подрядчики", payload="contractors_manage")],
        [CallbackButton(text="⬅️ Назад", payload="back_to_main")]
        ])


def get_add_menu(user_type: str):
    return inline_kb([
        [CallbackButton(text=f"Добавить {user_type}", payload=f"add_{user_type}")],
        [CallbackButton(text="⬅️ Назад", payload="back_to_manage")]
        ])


@router.message_created(Command("test"), IsAdmin())
async def cmd_test_admin(event: MessageCreated):
    sender = event.message.sender
    if sender is None:
        return
    # Текст ссылки max://user/... должен совпадать с именем в профиле MAX,
    # иначе клиент часто показывает сырой HTML вместо кликабельного профиля.
    admin_label = f"{sender.first_name} {(sender.last_name or '').strip()}".strip()
    text = fio_html(admin_label or sender.first_name, sender.user_id)
    await answer_message(event, text)


_DELETE_ROLE_ALIASES = {
    "manager": "manager",
    "менеджер": "manager",
    "security": "security",
    "sb": "security",
    "сб": "security",
    "resident": "resident",
    "резидент": "resident",
    "contractor": "contractor",
    "подрядчик": "contractor",
}


async def _admin_cmd_delete_resident(session, resident_id: int) -> bool:
    if not await session.get(Resident, resident_id):
        return False
    tp_ids = (
        await session.execute(
            select(TemporaryPass.id).where(TemporaryPass.resident_id == resident_id)
        )
    ).scalars().all()
    if tp_ids:
        await session.execute(
            delete(TempPassYooKassaPayment).where(
                TempPassYooKassaPayment.temporary_pass_id.in_(tp_ids)
            )
        )
    await session.execute(delete(TemporaryPass).where(TemporaryPass.resident_id == resident_id))
    await session.execute(delete(PermanentPass).where(PermanentPass.resident_id == resident_id))
    await session.execute(delete(Appeal).where(Appeal.resident_id == resident_id))
    await session.execute(
        delete(RegistrationRequest).where(RegistrationRequest.resident_id == resident_id)
    )
    await session.execute(
        delete(ResidentContractorRequest).where(
            ResidentContractorRequest.resident_id == resident_id
        )
    )
    await session.execute(delete(Resident).where(Resident.id == resident_id))
    return True


async def _admin_cmd_delete_contractor(session, contractor_id: int) -> bool:
    if not await session.get(Contractor, contractor_id):
        return False
    tp_ids = (
        await session.execute(
            select(TemporaryPass.id).where(TemporaryPass.contractor_id == contractor_id)
        )
    ).scalars().all()
    if tp_ids:
        await session.execute(
            delete(TempPassYooKassaPayment).where(
                TempPassYooKassaPayment.temporary_pass_id.in_(tp_ids)
            )
        )
    await session.execute(
        delete(TemporaryPass).where(TemporaryPass.contractor_id == contractor_id)
    )
    await session.execute(
        delete(ContractorRegistrationRequest).where(
            ContractorRegistrationRequest.contractor_id == contractor_id
        )
    )
    await session.execute(
        delete(ContractorContractorRequest).where(
            ContractorContractorRequest.contractor_id == contractor_id
        )
    )
    await session.execute(delete(Contractor).where(Contractor.id == contractor_id))
    return True


@router.message_created(Command("delete"), IsAdmin())
async def cmd_delete_user_by_id(event: MessageCreated):
    """Удаление менеджера / СБ / резидента / подрядчика по ID (только админ)."""
    text = text_from_message(event)
    if not text:
        return
    parts = text.split()
    if len(parts) < 3:
        await answer_message(
            event,
            "Использование: /delete <тип> <id>\n\n"
            "Типы: manager (менеджер), security (сб), resident (резидент), "
            "contractor (подрядчик)\n\n"
            "Пример: /delete manager 2",
        )
        return
    role_key = _DELETE_ROLE_ALIASES.get(parts[1].strip().lower())
    if role_key is None:
        await answer_message(
            event,
            f"Неизвестный тип «{parts[1]}». Укажите: manager, security, resident или contractor.",
        )
        return
    try:
        pk = int(parts[2])
    except ValueError:
        await answer_message(event, "Второй аргумент должен быть числовым ID.")
        return
    if pk < 1:
        await answer_message(event, "ID должен быть положительным числом.")
        return

    try:
        async with AsyncSessionLocal() as session:
            if role_key == "manager":
                row = await session.get(Manager, pk)
                if not row:
                    await answer_message(event, f"Менеджер с ID {pk} не найден.")
                    return
                await session.delete(row)
            elif role_key == "security":
                row = await session.get(Security, pk)
                if not row:
                    await answer_message(event, f"Сотрудник СБ с ID {pk} не найден.")
                    return
                await session.delete(row)
            elif role_key == "resident":
                ok = await _admin_cmd_delete_resident(session, pk)
                if not ok:
                    await answer_message(event, f"Резидент с ID {pk} не найден.")
                    return
            else:
                ok = await _admin_cmd_delete_contractor(session, pk)
                if not ok:
                    await answer_message(event, f"Подрядчик с ID {pk} не найден.")
                    return
            await session.commit()
        label = {
            "manager": "Менеджер",
            "security": "СБ",
            "resident": "Резидент",
            "contractor": "Подрядчик",
        }[role_key]
        await answer_message(event, f"✅ {label} id={pk} удалён из базы.")
    except Exception as e:
        await bot.send_message(user_id=RAZRAB, text=f"{user_id_from_message(event)} — /delete: {e!s}")
        await answer_message(event, f"❌ Ошибка при удалении: {e!s}")


@router.message_created(CommandStart())
async def process_start_admin(event: MessageCreated):
    try:
        uid = user_id_from_message(event)
        kb = (
            get_admin_menu()
            if uid is not None and uid in ADMIN_IDS
            else get_manager_menu()
        )
        await answer_message(
            event,
            text="Здравствуйте!\n\nДобро пожаловать в Главное меню",
            kb=kb,
        )
    except Exception as e:
        await bot.send_message(user_id=RAZRAB, text=f'{user_id_from_message(event)} - {str(e)}')
        await asyncio.sleep(0.05)


@router.message_created(F.message.body.text == "Главное меню")
async def main_menu(event: MessageCreated, context: BaseContext):
    try:
        await context.clear()
        uid = user_id_from_message(event)
        kb = (
            get_admin_menu()
            if uid is not None and uid in ADMIN_IDS
            else get_manager_menu()
        )
        await answer_message(
            event,
            text="Добро пожаловать в Главное меню",
            kb=kb,
        )
    except Exception as e:
        await bot.send_message(user_id=RAZRAB, text=f'{user_id_from_message(event)} - {str(e)}')
        await asyncio.sleep(0.05)


@router.message_callback(F.callback.payload.in_({"user_management", "back_to_manage"}))
async def user_management(event: MessageCallback):
    try:
        if event.callback.user.user_id in ADMIN_IDS:
            kb = get_admin_user_management_menu()
        else:
            kb = get_manager_user_management_menu()
        await edit_or_send_callback(bot, event, "Выберите категорию пользователей:", kb)
    except Exception as e:
        await bot.send_message(user_id=RAZRAB, text=f'{event.callback.user.user_id} - {str(e)}')
        await asyncio.sleep(0.05)


@router.message_callback(F.callback.payload.endswith("manage"))
async def manage_category(event: MessageCallback, context: BaseContext):
    try:
        user_type = event.callback.payload.split("_")[0]
        await context.update_data(user_type=user_type)

        # Для резидентов
        if user_type == 'residents':
            keyboard = inline_kb([
                [CallbackButton(text="Добавить резидента", payload=f"add_{user_type}")],
                [CallbackButton(text="Список резидентов", payload="list_residents")],
                [CallbackButton(text="⬅️ Назад", payload="back_to_manage")]
            ])
        # Для подрядчиков
        elif user_type == 'contractors':
            keyboard = inline_kb([
                [CallbackButton(text="Добавить подрядчика", payload=f"add_{user_type}")],
                [CallbackButton(text="Список подрядчиков", payload="list_contractors")],
                [CallbackButton(text="⬅️ Назад", payload="back_to_manage")]
            ])
        elif user_type == 'managers':
            keyboard = inline_kb([
                [CallbackButton(text="Добавить менеджера", payload=f"add_{user_type}")],
                [CallbackButton(text="Список менеджеров", payload="list_managers")],
                [CallbackButton(text="⬅️ Назад", payload="back_to_manage")]
            ])
            # Для СБ
        elif user_type == 'security':
            keyboard = inline_kb([
                [CallbackButton(text="Добавить СБ", payload=f"add_{user_type}")],
                [CallbackButton(text="Список СБ", payload="list_security")],
                [CallbackButton(text="⬅️ Назад", payload="back_to_manage")]
            ])
        else:
            return

        await edit_or_send_callback(bot, event, f"Управление {user_type}:", keyboard)
    except Exception as e:
        await bot.send_message(user_id=RAZRAB, text=f'{event.callback.user.user_id} - {str(e)}')
        await asyncio.sleep(0.05)


@router.message_callback(F.callback.payload.startswith("add_"))
async def start_add_user(event: MessageCallback, context: BaseContext):
    try:
        await context.set_state(AddUserStates.WAITING_PHONE)
        await event.message.answer("Введите телефон пользователя:")
    except Exception as e:
        await bot.send_message(user_id=RAZRAB, text=f'{event.callback.user.user_id} - {str(e)}')
        await asyncio.sleep(0.05)


@router.message_created(F.message.body.text, AddUserStates.WAITING_PHONE)
async def process_phone(event: MessageCreated, context: BaseContext):
    try:
        data = await context.get_data()
        user_type = data['user_type']
        phone = (text_from_message(event) or '')
        if not is_valid_phone(phone):
            await answer_message(event, 'Телефон должен быть в формате 8XXXXXXXXXX.\nПопробуйте ввести еще раз!')
            return

        async with AsyncSessionLocal() as session:
            try:
                if user_type == 'managers':
                    new_user = Manager(phone=phone)
                elif user_type == 'security':
                    new_user = Security(phone=phone)
                elif user_type == 'residents':
                    new_user = Resident(phone=phone)
                elif user_type == 'contractors':
                    new_user = Contractor(phone=phone)

                session.add(new_user)
                await session.commit()
                await answer_message(event, f"Пользователь с телефоном {phone} добавлен в {user_type}!")
                if user_type == 'residents':
                    keyboard = inline_kb([
                        [CallbackButton(text="Добавить резидента", payload=f"add_{user_type}")],
                        [CallbackButton(text="Список резидентов", payload="list_residents")],
                        [CallbackButton(text="⬅️ Назад", payload="back_to_manage")]
                    ])
                # Для подрядчиков
                elif user_type == 'contractors':
                    keyboard = inline_kb([
                        [CallbackButton(text="Добавить подрядчика", payload=f"add_{user_type}")],
                        [CallbackButton(text="Список подрядчиков", payload="list_contractors")],
                        [CallbackButton(text="⬅️ Назад", payload="back_to_manage")]
                    ])
                # Для остальных
                elif user_type == 'managers':
                    keyboard = inline_kb([
                        [CallbackButton(text="Добавить менеджера", payload=f"add_{user_type}")],
                        [CallbackButton(text="Список менеджеров", payload="list_managers")],
                        [CallbackButton(text="⬅️ Назад", payload="back_to_manage")]
                    ])
                    # Для СБ
                elif user_type == 'security':
                    keyboard = inline_kb([
                        [CallbackButton(text="Добавить СБ", payload=f"add_{user_type}")],
                        [CallbackButton(text="Список СБ", payload="list_security")],
                        [CallbackButton(text="⬅️ Назад", payload="back_to_manage")]
                    ])
                else:
                    return

                await answer_message(event, 
                    text=f"Управление {user_type}:",
                    kb=keyboard
                )

            except Exception as e:
                await answer_message(event, f"Ошибка: {str(e)}")
                await session.rollback()

    except Exception as e:
        await bot.send_message(user_id=RAZRAB, text=f'{user_id_from_message(event)} - {str(e)}')
        await asyncio.sleep(0.05)


@router.message_callback(F.callback.payload.startswith("list_residents"))
async def show_residents_list(event: MessageCallback):
    try:
        page = _parse_list_page(event.callback.payload, "list_residents")
        if page is None:
            return

        async with AsyncSessionLocal() as session:
            total = await session.scalar(
                select(func.count(Resident.id)).where(Resident.status.is_(True))
            )
            total = int(total or 0)

            if total == 0:
                await callback_ack(bot, event, "Нет зарегистрированных резидентов")
                return

            max_page = max(0, (total - 1) // PAGE_SIZE)
            page = max(0, min(page, max_page))

            result = await session.execute(
                select(Resident)
                .where(Resident.status.is_(True))
                .order_by(Resident.id)
                .offset(page * PAGE_SIZE)
                .limit(PAGE_SIZE)
            )
            residents = result.scalars().all()

            buttons = []
            for resident in residents:
                buttons.append(
                    [
                        CallbackButton(
                            text=_resident_list_button_text(resident),
                            payload=f"view_resident_{resident.id}",
                        )
                    ]
                )

            nav: list[CallbackButton] = []
            if page > 0:
                nav.append(
                    CallbackButton(
                        text="⬅️ Предыдущие",
                        payload=f"list_residents_p_{page - 1}",
                    )
                )
            if (page + 1) * PAGE_SIZE < total:
                nav.append(
                    CallbackButton(
                        text="Следующие ➡️",
                        payload=f"list_residents_p_{page + 1}",
                    )
                )
            if nav:
                buttons.append(nav)

            buttons.append([CallbackButton(text="⬅️ Назад", payload="residents_manage")])

            header = (
                f"Список зарегистрированных резидентов "
                f"(стр. {page + 1}/{max_page + 1}, всего {total}):"
            )
            try:
                await edit_or_send_callback(bot, event, header, inline_kb(buttons))
            except Exception:
                await event.message.answer(
                    text=header,
                    attachments=[inline_kb(buttons).as_markup()],
                )
    except Exception as e:
        await bot.send_message(user_id=RAZRAB, text=f'{event.callback.user.user_id} - {str(e)}')
        await asyncio.sleep(0.05)


@router.message_callback(F.callback.payload.startswith("view_resident_"))
async def view_resident_details(event: MessageCallback):
    try:
        resident_id = int(event.callback.payload.split("_")[-1])
        async with AsyncSessionLocal() as session:
            resident = await session.get(Resident, resident_id)
            if not resident:
                await callback_ack(bot, event, "Резидент не найден")
                return

            # Формируем текст
            text = (
                f"ID: {resident.id}\n"
                f"ФИО: {fio_html(resident.fio, resident.tg_id)}\n"
                f"{profile_link_line_html(resident.first_name, resident.last_name, resident.tg_id, fallback_fio=resident.fio)}"
                f"Телефон: {html_lib.escape(str(resident.phone or ''))}\n"
                f"Номер участка: {html_lib.escape(str(resident.plot_number or ''))}\n"
                f"Время регистрации: {resident.time_registration}"
            )

            # Клавиатура с кнопкой "Назад" к списку резидентов
            keyboard = inline_kb([
                [CallbackButton(text="🗑 Удалить", payload=f"delete_resident_{resident_id}")],
                [CallbackButton(text="⬅️ Назад", payload="list_residents")]
            ])

            await edit_or_send_callback(bot, event, text, keyboard, parse_mode=ParseMode.HTML)
            await callback_ack(bot, event)
    except Exception as e:
        await bot.send_message(user_id=RAZRAB, text=f'{event.callback.user.user_id} - {str(e)}')
        await asyncio.sleep(0.05)


@router.message_callback(F.callback.payload.startswith("list_contractors"))
async def show_contractors_list(event: MessageCallback):
    try:
        page = _parse_list_page(event.callback.payload, "list_contractors")
        if page is None:
            return

        async with AsyncSessionLocal() as session:
            total = await session.scalar(
                select(func.count(Contractor.id)).where(Contractor.status.is_(True))
            )
            total = int(total or 0)

            if total == 0:
                await callback_ack(bot, event, "Нет зарегистрированных подрядчиков")
                return

            max_page = max(0, (total - 1) // PAGE_SIZE)
            page = max(0, min(page, max_page))

            result = await session.execute(
                select(Contractor)
                .where(Contractor.status.is_(True))
                .order_by(Contractor.id)
                .offset(page * PAGE_SIZE)
                .limit(PAGE_SIZE)
            )
            contractors = result.scalars().all()

            buttons = []
            for contractor in contractors:
                buttons.append(
                    [
                        CallbackButton(
                            text=_contractor_list_button_text(contractor),
                            payload=f"view_contractor_{contractor.id}",
                        )
                    ]
                )

            nav: list[CallbackButton] = []
            if page > 0:
                nav.append(
                    CallbackButton(
                        text="⬅️ Предыдущие",
                        payload=f"list_contractors_p_{page - 1}",
                    )
                )
            if (page + 1) * PAGE_SIZE < total:
                nav.append(
                    CallbackButton(
                        text="Следующие ➡️",
                        payload=f"list_contractors_p_{page + 1}",
                    )
                )
            if nav:
                buttons.append(nav)

            buttons.append([CallbackButton(text="⬅️ Назад", payload="contractors_manage")])

            header = (
                f"Список зарегистрированных подрядчиков "
                f"(стр. {page + 1}/{max_page + 1}, всего {total}):"
            )
            try:
                await edit_or_send_callback(bot, event, header, inline_kb(buttons))
            except Exception:
                await event.message.answer(
                    text=header,
                    attachments=[inline_kb(buttons).as_markup()],
                )
    except Exception as e:
        await bot.send_message(user_id=RAZRAB, text=f'{event.callback.user.user_id} - {str(e)}')
        await asyncio.sleep(0.05)


@router.message_callback(F.callback.payload.startswith("view_contractor_"))
async def view_contractor_details(event: MessageCallback):
    try:
        contractor_id = int(event.callback.payload.split("_")[-1])
        async with AsyncSessionLocal() as session:
            contractor = await session.get(Contractor, contractor_id)
            if not contractor:
                await callback_ack(bot, event, "Подрядчик не найден")
                return

            # Формируем текст
            text = (
                f"ID: {contractor.id}\n"
                f"ФИО: {fio_html(contractor.fio, contractor.tg_id)}\n"
                f"{profile_link_line_html(contractor.first_name, contractor.last_name, contractor.tg_id, fallback_fio=contractor.fio)}"
                f"Телефон: {html_lib.escape(str(contractor.phone or ''))}\n"
                f"Компания: {html_lib.escape(str(contractor.company or ''))}\n"
                f"Должность: {html_lib.escape(str(contractor.position or ''))}\n"
                f"Принадлежность: {html_lib.escape(str(contractor.affiliation or ''))}\n"
                f"Возможность добавлять субподрядчиков: {contractor.can_add_contractor}\n"
                f"Время регистрации: {contractor.time_registration}"
            )
            if contractor.can_add_contractor == True:
                text_admin = '✅Подрядчик-администратор'
            else:
                text_admin = '❌Подрядчик-администратор'

            # Клавиатура с кнопкой "Назад" к списку подрядчиков
            keyboard = inline_kb([
                [CallbackButton(text="🗑 Удалить", payload=f"delete_contractor_{contractor_id}")],
                [CallbackButton(text=text_admin, payload=f"change_admin_{contractor_id}")],
                [CallbackButton(text="⬅️ Назад", payload="list_contractors")]
            ])

            await edit_or_send_callback(bot, event, text, keyboard, parse_mode=ParseMode.HTML)
            await callback_ack(bot, event)
    except Exception as e:
        await bot.send_message(user_id=RAZRAB, text=f'{event.callback.user.user_id} - {str(e)}')
        await asyncio.sleep(0.05)


@router.message_callback(F.callback.payload.startswith("delete_resident_"))
async def confirm_delete_resident(event: MessageCallback):
    try:
        resident_id = int(event.callback.payload.split("_")[-1])
        await event.message.answer(
            text="Вы точно хотите удалить резидента?",
            attachments=[inline_kb([
                [CallbackButton(text="✅ Да", payload=f"confirm_delete_yes_{resident_id}")],
                [CallbackButton(text="❌ Нет", payload=f"confirm_delete_no_{resident_id}")]
            ]).as_markup()],
        )
    except Exception as e:
        await bot.send_message(user_id=RAZRAB, text=f'{event.callback.user.user_id} - {str(e)}')
        await asyncio.sleep(0.05)


@router.message_callback(F.callback.payload.startswith("confirm_delete_no_"))
async def cancel_delete(event: MessageCallback):
    try:
        resident_id = int(event.callback.payload.split("_")[-1])
        # Возвращаемся к просмотру резидента
        await view_resident_details(callback)
    except Exception as e:
        await bot.send_message(user_id=RAZRAB, text=f'{event.callback.user.user_id} - {str(e)}')
        await asyncio.sleep(0.05)


@router.message_callback(F.callback.payload.startswith("confirm_delete_yes_"))
async def execute_delete(event: MessageCallback, context: BaseContext):
    try:
        resident_id = int(event.callback.payload.split("_")[-1])
        async with AsyncSessionLocal() as session:
            # Удаляем связанные заявки
            stmt1 = delete(RegistrationRequest).where(RegistrationRequest.resident_id == resident_id)
            stmt2 = delete(ResidentContractorRequest).where(ResidentContractorRequest.resident_id == resident_id)
            stmt3 = delete(PermanentPass).where(PermanentPass.resident_id == resident_id)
            stmt4 = delete(TemporaryPass).where(TemporaryPass.resident_id == resident_id)
            stmt5 = delete(Appeal).where(Appeal.resident_id == resident_id)
            await session.execute(stmt1)
            await session.execute(stmt2)
            await session.execute(stmt3)
            await session.execute(stmt4)
            await session.execute(stmt5)
            resident = await session.get(Resident, resident_id)
            await bot.send_message(
                user_id=resident.tg_id,
                text='Вам ограничили доступ, если это случилось по ошибке обратитесь в управляющую компанию "Ели Estate"',
            )
            # Удаляем резидента
            stmt6 = delete(Resident).where(Resident.id == resident_id)
            await session.execute(stmt6)
            await session.commit()

        await event.message.answer("✅ Резидент удален")
        # Возвращаемся в меню управления резидентами
        keyboard = inline_kb([
            [CallbackButton(text="Добавить резидента", payload=f"add_residents")],
            [CallbackButton(text="Список резидентов", payload="list_residents")],
            [CallbackButton(text="⬅️ Назад", payload="back_to_manage")]
        ])
        await event.message.answer(
            text=f"Управление residents:",
            attachments=[keyboard.as_markup()],
        )
    except Exception as e:
        await bot.send_message(user_id=RAZRAB, text=f'{event.callback.user.user_id} - {str(e)}')
        await asyncio.sleep(0.05)


# Подтверждение удаления подрядчика
@router.message_callback(F.callback.payload.startswith("delete_contractor_"))
async def confirm_delete_contractor(event: MessageCallback):
    try:
        contractor_id = int(event.callback.payload.split("_")[-1])
        await event.message.answer(
            text="Вы точно хотите удалить подрядчика?",
            attachments=[inline_kb([
                [CallbackButton(text="✅ Да", payload=f"confirm_del_cont_yes_{contractor_id}")],
                [CallbackButton(text="❌ Нет", payload=f"confirm_del_cont_no_{contractor_id}")]
            ]).as_markup()],
        )
    except Exception as e:
        await bot.send_message(user_id=RAZRAB, text=f'{event.callback.user.user_id} - {str(e)}')
        await asyncio.sleep(0.05)


# Отмена удаления подрядчика
@router.message_callback(F.callback.payload.startswith("confirm_del_cont_no_"))
async def cancel_delete_contractor(event: MessageCallback):
    try:
        contractor_id = int(event.callback.payload.split("_")[-1])
        # Возвращаемся к просмотру подрядчика
        await view_contractor_details(callback)
    except Exception as e:
        await bot.send_message(user_id=RAZRAB, text=f'{event.callback.user.user_id} - {str(e)}')
        await asyncio.sleep(0.05)


# Выполнение удаления подрядчика
@router.message_callback(F.callback.payload.startswith("confirm_del_cont_yes_"))
async def execute_delete_contractor(event: MessageCallback, context: BaseContext):
    try:
        contractor_id = int(event.callback.payload.split("_")[-1])
        async with AsyncSessionLocal() as session:
            # Удаляем связанные записи
            stmt1 = delete(ContractorRegistrationRequest).where(
                ContractorRegistrationRequest.contractor_id == contractor_id
            )
            stmt2 = delete(TemporaryPass).where(
                TemporaryPass.contractor_id == contractor_id
            )
            await session.execute(stmt1)
            await session.execute(stmt2)
            contractor = await session.get(Contractor, contractor_id)
            await bot.send_message(
                user_id=contractor.tg_id,
                text='Вам ограничили доступ, если это случилось по ошибке обратитесь в управляющую компанию "Ели Estate"',
            )
            # Удаляем самого подрядчика
            stmt3 = delete(Contractor).where(Contractor.id == contractor_id)
            await session.execute(stmt3)
            await session.commit()

        await event.message.answer("✅ Подрядчик удален")

        # Возвращаемся к списку подрядчиков
        keyboard = inline_kb([
            [CallbackButton(text="Добавить подрядчика", payload=f"add_contractors")],
            [CallbackButton(text="Список подрядчиков", payload="list_contractors")],
            [CallbackButton(text="⬅️ Назад", payload="back_to_manage")]
        ])
        await event.message.answer(
            text=f"Управление contractors:",
            attachments=[keyboard.as_markup()],
        )
    except Exception as e:
        await bot.send_message(user_id=RAZRAB, text=f'{event.callback.user.user_id} - {str(e)}')
        await asyncio.sleep(0.05)


@router.message_callback(F.callback.payload == "list_managers")
async def show_managers_list(event: MessageCallback):
    try:
        async with AsyncSessionLocal() as session:
            result = await session.execute(
                select(Manager).where(Manager.status == True))
            managers = result.scalars().all()

            if not managers:
                await callback_ack(bot, event, "Нет зарегистрированных менеджеров")
                return

            buttons = []
            for manager in managers:
                button_text = f"{manager.fio}"
                if len(button_text) > 30:
                    button_text = button_text[:27] + "..."

                buttons.append([CallbackButton(
                    text=button_text,
                    payload=f"view_manager_{manager.id}"
                )])

            buttons.append([CallbackButton(
                text="⬅️ Назад",
                payload="managers_manage"
            )])

            await edit_or_send_callback(bot, event, "Список менеджеров:", inline_kb(buttons))
    except Exception as e:
        await bot.send_message(user_id=RAZRAB, text=f'{event.callback.user.user_id} - {str(e)}')
        await asyncio.sleep(0.05)


@router.message_callback(F.callback.payload == "list_security")
async def show_security_list(event: MessageCallback):
    try:
        async with AsyncSessionLocal() as session:
            result = await session.execute(
                select(Security).where(Security.status == True))
            security_list = result.scalars().all()

            if not security_list:
                await callback_ack(bot, event, "Нет зарегистрированных сотрудников СБ")
                return

            buttons = []
            for security in security_list:
                button_text = f"{security.fio}"
                if len(button_text) > 30:
                    button_text = button_text[:27] + "..."

                buttons.append([CallbackButton(
                    text=button_text,
                    payload=f"view_security_{security.id}"
                )])

            buttons.append([CallbackButton(
                text="⬅️ Назад",
                payload="security_manage"
            )])

            await edit_or_send_callback(bot, event, "Список сотрудников СБ:", inline_kb(buttons))
    except Exception as e:
        await bot.send_message(user_id=RAZRAB, text=f'{event.callback.user.user_id} - {str(e)}')
        await asyncio.sleep(0.05)


@router.message_callback(F.callback.payload.startswith("view_manager_"))
async def view_manager_details(event: MessageCallback):
    try:
        manager_id = int(event.callback.payload.split("_")[-1])
        async with AsyncSessionLocal() as session:
            manager = await session.get(Manager, manager_id)
            if not manager:
                await callback_ack(bot, event, "Менеджер не найден")
                return

            text = (
                f"ID: {manager.id}\n"
                f"ФИО: {fio_html(manager.fio, manager.tg_id)}\n"
                f"{profile_link_line_html(manager.first_name, manager.last_name, manager.tg_id, fallback_fio=manager.fio)}"
                f"Телефон: {html_lib.escape(str(manager.phone or ''))}\n"
                f"Время добавления: {manager.time_add_to_db}\n"
                f"Время регистрации: {manager.time_registration}\n"
                f"Статус: {'Активен' if manager.status else 'Неактивен'}"
            )

            keyboard = inline_kb([
                [CallbackButton(text="🗑 Удалить", payload=f"delete_manager_{manager_id}")],
                [CallbackButton(text="⬅️ Назад", payload="list_managers")]
            ])

            await edit_or_send_callback(bot, event, text, keyboard, parse_mode=ParseMode.HTML)
            await callback_ack(bot, event)
    except Exception as e:
        await bot.send_message(user_id=RAZRAB, text=f'{event.callback.user.user_id} - {str(e)}')
        await asyncio.sleep(0.05)

@router.message_callback(F.callback.payload.startswith("view_security_"))
async def view_security_details(event: MessageCallback):
    try:
        security_id = int(event.callback.payload.split("_")[-1])
        async with AsyncSessionLocal() as session:
            security = await session.get(Security, security_id)
            if not security:
                await callback_ack(bot, event, "Сотрудник СБ не найден")
                return

            text = (
                f"ID: {security.id}\n"
                f"ФИО: {fio_html(security.fio, security.tg_id)}\n"
                f"{profile_link_line_html(security.first_name, security.last_name, security.tg_id, fallback_fio=security.fio)}"
                f"Телефон: {html_lib.escape(str(security.phone or ''))}\n"
                f"Время добавления: {security.time_add_to_db}\n"
                f"Время регистрации: {security.time_registration}\n"
                f"Статус: {'Активен' if security.status else 'Неактивен'}"
            )

            keyboard = inline_kb([
                [CallbackButton(text="🗑 Удалить", payload=f"delete_security_{security_id}")],
                [CallbackButton(text="⬅️ Назад", payload="list_security")]
            ])

            await edit_or_send_callback(bot, event, text, keyboard, parse_mode=ParseMode.HTML)
            await callback_ack(bot, event)
    except Exception as e:
        await bot.send_message(user_id=RAZRAB, text=f'{event.callback.user.user_id} - {str(e)}')
        await asyncio.sleep(0.05)


@router.message_callback(F.callback.payload.startswith("delete_manager_"))
async def confirm_delete_manager(event: MessageCallback):
    try:
        manager_id = int(event.callback.payload.split("_")[-1])
        await edit_or_send_callback(bot, event, "Вы точно хотите удалить менеджера?", inline_kb([
                [CallbackButton(text="✅ Да", payload=f"confirm_delete_manager_yes_{manager_id}")],
                [CallbackButton(text="❌ Нет", payload=f"confirm_delete_manager_no_{manager_id}")]
            ]))
    except Exception as e:
        await bot.send_message(user_id=RAZRAB, text=f'{event.callback.user.user_id} - {str(e)}')
        await asyncio.sleep(0.05)

@router.message_callback(F.callback.payload.startswith("delete_security_"))
async def confirm_delete_security(event: MessageCallback):
    try:
        security_id = int(event.callback.payload.split("_")[-1])
        await edit_or_send_callback(bot, event, "Вы точно хотите удалить сотрудника СБ?", inline_kb([
                [CallbackButton(text="✅ Да", payload=f"confirm_delete_security_yes_{security_id}")],
                [CallbackButton(text="❌ Нет", payload=f"confirm_delete_security_no_{security_id}")]
            ]))
    except Exception as e:
        await bot.send_message(user_id=RAZRAB, text=f'{event.callback.user.user_id} - {str(e)}')
        await asyncio.sleep(0.05)


@router.message_callback(F.callback.payload.startswith("confirm_delete_manager_yes_"))
async def execute_delete_manager(event: MessageCallback, context: BaseContext):
    try:
        manager_id = int(event.callback.payload.split("_")[-1])
        async with AsyncSessionLocal() as session:
            manager = await session.get(Manager, manager_id)
            await bot.send_message(
                user_id=manager.tg_id,
                text='Вам ограничили доступ, если это случилось по ошибке обратитесь в управляющую компанию "Ели Estate"',
            )
            stmt = delete(Manager).where(Manager.id == manager_id)
            await session.execute(stmt)
            await session.commit()

        await event.message.answer("✅ Менеджер удален")

        # Возвращаемся в меню управления менеджерами
        keyboard = inline_kb([
            [CallbackButton(text="Добавить менеджера", payload=f"add_managers")],
            [CallbackButton(text="Список менеджеров", payload="list_managers")],
            [CallbackButton(text="⬅️ Назад", payload="back_to_manage")]
        ])

        await event.message.answer(
            text=f"Управление managers:",
            attachments=[keyboard.as_markup()],
        )
    except Exception as e:
        await bot.send_message(user_id=RAZRAB, text=f'{event.callback.user.user_id} - {str(e)}')
        await asyncio.sleep(0.05)


@router.message_callback(F.callback.payload.startswith("confirm_delete_security_yes_"))
async def execute_delete_security(event: MessageCallback, context: BaseContext):
    try:
        security_id = int(event.callback.payload.split("_")[-1])
        async with AsyncSessionLocal() as session:
            security = await session.get(Security, security_id)
            await bot.send_message(
                user_id=security.tg_id,
                text='Вам ограничили доступ, если это случилось по ошибке обратитесь в управляющую компанию "Ели Estate"',
            )
            stmt = delete(Security).where(Security.id == security_id)
            await session.execute(stmt)
            await session.commit()

        await event.message.answer("✅ Сотрудник СБ удален")

        # Возвращаемся в меню управления СБ
        keyboard = inline_kb([
            [CallbackButton(text="Добавить СБ", payload=f"add_security")],
            [CallbackButton(text="Список СБ", payload="list_security")],
            [CallbackButton(text="⬅️ Назад", payload="back_to_manage")]
        ])

        await event.message.answer(
            text=f"Управление security",
            attachments=[keyboard.as_markup()],
        )
    except Exception as e:
        await bot.send_message(user_id=RAZRAB, text=f'{event.callback.user.user_id} - {str(e)}')
        await asyncio.sleep(0.05)


@router.message_callback(F.callback.payload.startswith("confirm_delete_manager_no_"))
async def execute_no_delete_manager(event: MessageCallback, context: BaseContext):
    try:
        manager_id = int(event.callback.payload.split("_")[-1])
        async with AsyncSessionLocal() as session:
            manager = await session.get(Manager, manager_id)
            if not manager:
                await callback_ack(bot, event, "Менеджер не найден")
                return

            text = (
                f"ID: {manager.id}\n"
                f"ФИО: {fio_html(manager.fio, manager.tg_id)}\n"
                f"{profile_link_line_html(manager.first_name, manager.last_name, manager.tg_id, fallback_fio=manager.fio)}"
                f"Телефон: {html_lib.escape(str(manager.phone or ''))}\n"
                f"Время добавления: {manager.time_add_to_db}\n"
                f"Время регистрации: {manager.time_registration}\n"
                f"Статус: {'Активен' if manager.status else 'Неактивен'}"
            )

            keyboard = inline_kb([
                [CallbackButton(text="🗑 Удалить", payload=f"delete_manager_{manager_id}")],
                [CallbackButton(text="⬅️ Назад", payload="list_managers")]
            ])

            await edit_or_send_callback(bot, event, text, keyboard, parse_mode=ParseMode.HTML)
            await callback_ack(bot, event)
    except Exception as e:
        await bot.send_message(user_id=RAZRAB, text=f'{event.callback.user.user_id} - {str(e)}')
        await asyncio.sleep(0.05)


@router.message_callback(F.callback.payload.startswith("confirm_delete_security_no_"))
async def execute_no_delete_security(event: MessageCallback, context: BaseContext):
    try:
        security_id = int(event.callback.payload.split("_")[-1])
        async with AsyncSessionLocal() as session:
            security = await session.get(Security, security_id)
            if not security:
                await callback_ack(bot, event, "Сотрудник СБ не найден")
                return

            text = (
                f"ID: {security.id}\n"
                f"ФИО: {fio_html(security.fio, security.tg_id)}\n"
                f"{profile_link_line_html(security.first_name, security.last_name, security.tg_id, fallback_fio=security.fio)}"
                f"Телефон: {html_lib.escape(str(security.phone or ''))}\n"
                f"Время добавления: {security.time_add_to_db}\n"
                f"Время регистрации: {security.time_registration}\n"
                f"Статус: {'Активен' if security.status else 'Неактивен'}"
            )

            keyboard = inline_kb([
                [CallbackButton(text="🗑 Удалить", payload=f"delete_security_{security_id}")],
                [CallbackButton(text="⬅️ Назад", payload="list_security")]
            ])

            await edit_or_send_callback(bot, event, text, keyboard, parse_mode=ParseMode.HTML)
            await callback_ack(bot, event)
    except Exception as e:
        await bot.send_message(user_id=RAZRAB, text=f'{event.callback.user.user_id} - {str(e)}')
        await asyncio.sleep(0.05)


@router.message_callback(F.callback.payload.startswith("change_admin_"))
async def change_contractor_admin(event: MessageCallback):
    try:
        contractor_id = int(event.callback.payload.split("_")[-1])
        async with AsyncSessionLocal() as session:
            contractor = await session.get(Contractor, contractor_id)
            if not contractor:
                await callback_ack(bot, event, "Подрядчик не найден")
                return

            # Формируем текст
            text = (
                f"ID: {contractor.id}\n"
                f"ФИО: {fio_html(contractor.fio, contractor.tg_id)}\n"
                f"{profile_link_line_html(contractor.first_name, contractor.last_name, contractor.tg_id, fallback_fio=contractor.fio)}"
                f"Телефон: {html_lib.escape(str(contractor.phone or ''))}\n"
                f"Компания: {html_lib.escape(str(contractor.company or ''))}\n"
                f"Должность: {html_lib.escape(str(contractor.position or ''))}\n"
                f"Принадлежность: {html_lib.escape(str(contractor.affiliation or ''))}\n"
                f"Возможность добавлять субподрядчиков: {not contractor.can_add_contractor}\n"
                f"Время регистрации: {contractor.time_registration}"
            )
            if contractor.can_add_contractor == False:
                text_admin = '✅Подрядчик-администратор'
            else:
                text_admin = '❌Подрядчик-администратор'

            # Клавиатура с кнопкой "Назад" к списку подрядчиков
            keyboard = inline_kb([
                [CallbackButton(text="🗑 Удалить", payload=f"delete_contractor_{contractor_id}")],
                [CallbackButton(text=text_admin, payload=f"change_admin_{contractor_id}")],
                [CallbackButton(text="⬅️ Назад", payload="list_contractors")]
            ])
            await edit_or_send_callback(bot, event, text, keyboard, parse_mode=ParseMode.HTML)
            await callback_ack(bot, event)
            contractor.can_add_contractor = not contractor.can_add_contractor
            await session.commit()
    except Exception as e:
        await bot.send_message(user_id=RAZRAB, text=f'{event.callback.user.user_id} - {str(e)}')
        await asyncio.sleep(0.05)


@router.message_created(Command("import"))
async def command_export(event: MessageCreated, context: BaseContext):
    try:
        await context.clear()
        await answer_message(event, "Пожалуйста, загрузите Excel-файл с резидентами.")
        await context.set_state(ExportStates.WAITING_FILE)
    except Exception as e:
        await bot.send_message(user_id=RAZRAB, text=f'{user_id_from_message(event)} - {str(e)}')
        await asyncio.sleep(0.05)


@router.message_created(ExportStates.WAITING_FILE)
async def handle_export_file(event: MessageCreated, context: BaseContext) -> None:
    excel_file = None
    try:
        if not message.document.file_name.endswith('.xlsx'):
            await answer_message(event, "Пожалуйста, загрузите файл в формате xlsx")
            return

        file_id = message.document.file_id
        file = await bot.get_file(file_id)
        file_path = file.file_path

        excel_file = BytesIO()
        await bot.download_file(file_path, excel_file)
        excel_file.seek(0)

        wb = openpyxl.load_workbook(excel_file)
        sheet = wb.active
        errors = []
        success_count = 0
        total_rows = sheet.max_row

        async with AsyncSessionLocal() as session:
            for row in range(1, total_rows + 1):
                phone_cell = sheet.cell(row=row, column=1)
                fio_cell = sheet.cell(row=row, column=2)
                phone = str(phone_cell.value) if phone_cell.value is not None else ""
                fio = fio_cell.value

                if not phone or not is_valid_phone(phone) or not fio:
                    errors.append(f"Строка №{row} - не корректный телефон или фио должно быть заполнено")
                    continue

                try:
                    resident = Resident(phone=phone, fio=fio)
                    session.add(resident)
                    success_count += 1
                except Exception as e:
                    errors.append(f"Строка №{row} - ошибка при добавлении в базу: {str(e)}")

            await session.commit()

        report = f"Загружено {success_count} резидентов из {total_rows} строк."
        if errors:
            error_report = "\n".join(errors)
            report += f"\nОшибки:\n{error_report}"

        if len(report) > 4096:
            for x in range(0, len(report), 4096):
                await answer_message(event, report[x:x+4096])
        else:
            await answer_message(event, report)

    except Exception as e:
        await bot.send_message(user_id=RAZRAB, text=f'{user_id_from_message(event)} - {str(e)}')
        await asyncio.sleep(0.05)
    finally:
        if excel_file:
            excel_file.close()
        await context.clear()

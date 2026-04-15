# handlers_admin_statistic.py
import asyncio
from io import BytesIO

import openpyxl
from maxapi import F, Router
from maxapi.enums.parse_mode import ParseMode
from maxapi.enums.upload_type import UploadType
from maxapi.types.attachments.buttons.callback_button import CallbackButton
from maxapi.types.input_media import InputMediaBuffer
from maxapi.types.updates.message_callback import MessageCallback
from maxapi.utils.inline_keyboard import InlineKeyboardBuilder
from openpyxl.workbook import Workbook
from sqlalchemy import func, select

from bot import bot
from config import RAZRAB
from db.models import (
    AsyncSessionLocal,
    Contractor,
    Manager,
    PermanentPass,
    Resident,
    Security,
    TemporaryPass,
)
from filters import IsAdminOrManager
from max_helpers import callback_ack, edit_or_send_callback

router = Router(router_id="admin_statistic")
router.filter(IsAdminOrManager())


def _statistics_kb() -> InlineKeyboardBuilder:
    b = InlineKeyboardBuilder()
    b.row(CallbackButton(text="📤 Экспорт в xlsx", payload="export_to_xlsx"))
    b.row(CallbackButton(text="⬅️ Назад", payload="back_to_main"))
    return b


@router.message_callback(F.callback.payload == "statistics_menu")
async def show_statistics(event: MessageCallback) -> None:
    try:
        async with AsyncSessionLocal() as session:
            total_residents = await session.scalar(select(func.count(Resident.id)))
            registered_residents = await session.scalar(
                select(func.count(Resident.id)).where(Resident.status == True)  # noqa: E712
            )
            unregistered_residents = (total_residents or 0) - (registered_residents or 0)

            total_contractors = await session.scalar(select(func.count(Contractor.id)))
            registered_contractors = await session.scalar(
                select(func.count(Contractor.id)).where(Contractor.status == True)  # noqa: E712
            )
            unregistered_contractors = (total_contractors or 0) - (registered_contractors or 0)

            total_permanent = await session.scalar(select(func.count(PermanentPass.id)))
            pending_permanent = await session.scalar(
                select(func.count(PermanentPass.id)).where(PermanentPass.status == "pending")
            )
            approved_permanent = await session.scalar(
                select(func.count(PermanentPass.id)).where(PermanentPass.status == "approved")
            )
            rejected_permanent = await session.scalar(
                select(func.count(PermanentPass.id)).where(PermanentPass.status == "rejected")
            )

            total_temporary = await session.scalar(select(func.count(TemporaryPass.id)))
            pending_temporary = await session.scalar(
                select(func.count(TemporaryPass.id)).where(TemporaryPass.status == "pending")
            )
            approved_temporary = await session.scalar(
                select(func.count(TemporaryPass.id)).where(TemporaryPass.status == "approved")
            )
            rejected_temporary = await session.scalar(
                select(func.count(TemporaryPass.id)).where(TemporaryPass.status == "rejected")
            )

            total_passes = (total_permanent or 0) + (total_temporary or 0)
            pending_passes = (pending_permanent or 0) + (pending_temporary or 0)
            approved_passes = (approved_permanent or 0) + (approved_temporary or 0)
            rejected_passes = (rejected_temporary or 0) + (rejected_permanent or 0)

        text = (
            "📊 <b>Статистика системы</b>\n\n"
            "👤 <b>Резиденты:</b>\n"
            f"  Всего: {total_residents}\n"
            f"  Зарегистрированных: {registered_residents}\n"
            f"  Не зарегистрированных: {unregistered_residents}\n\n"
            "👷 <b>Подрядчики:</b>\n"
            f"  Всего: {total_contractors}\n"
            f"  Зарегистрированных: {registered_contractors}\n"
            f"  Не зарегистрированных: {unregistered_contractors}\n\n"
            "🎫 <b>Все пропуска:</b>\n"
            f"  Всего заявок: {total_passes}\n"
            f"  На утверждении: {pending_passes}\n"
            f"  Утвержденных: {approved_passes}\n"
            f"  Отклоненных: {rejected_passes}\n\n"
            "🔖 <b>Постоянные пропуска:</b>\n"
            f"  Всего заявок: {total_permanent}\n"
            f"  На утверждении: {pending_permanent}\n"
            f"  Утвержденных: {approved_permanent}\n"
            f"  Отклоненных: {rejected_permanent}\n\n"
            "⏳ <b>Временные пропуска:</b>\n"
            f"  Всего заявок: {total_temporary}\n"
            f"  На утверждении: {pending_temporary}\n"
            f"  Утвержденных: {approved_temporary}\n"
            f"  Отклоненных: {rejected_temporary}"
        )
        await edit_or_send_callback(
            bot, event, text, _statistics_kb(), parse_mode=ParseMode.HTML
        )
    except Exception as e:
        await bot.send_message(user_id=RAZRAB, text=f"{event.callback.user.user_id} - {e!s}")
        await asyncio.sleep(0.05)


@router.message_callback(F.callback.payload == "export_to_xlsx")
async def export_statistics_to_xlsx(event: MessageCallback) -> None:
    try:
        await callback_ack(bot, event, "Формируем отчет...")
        async with AsyncSessionLocal() as session:
            wb = Workbook()

            ws_res = wb.active
            ws_res.title = "Резиденты"
            residents = await session.execute(select(Resident))
            ws_res.append(["ID", "Телефон", "ФИО", "Участок", "TG ID", "Статус"])
            for res in residents.scalars():
                ws_res.append(
                    [
                        res.id,
                        res.phone,
                        res.fio,
                        res.plot_number,
                        res.tg_id,
                        "Активен" if res.status else "Неактивен",
                    ]
                )

            ws_contr = wb.create_sheet("Подрядчики")
            contractors = await session.execute(select(Contractor))
            ws_contr.append(
                ["ID", "Телефон", "ФИО", "Компания", "Должность", "TG ID", "Статус"]
            )
            for contr in contractors.scalars():
                ws_contr.append(
                    [
                        contr.id,
                        contr.phone,
                        contr.fio,
                        contr.company,
                        contr.position,
                        contr.tg_id,
                        "Активен" if contr.status else "Неактивен",
                    ]
                )

            ws_mgr = wb.create_sheet("Менеджеры")
            managers = await session.execute(select(Manager))
            ws_mgr.append(["ID", "Телефон", "ФИО", "TG ID", "Username", "Статус"])
            for mgr in managers.scalars():
                ws_mgr.append(
                    [
                        mgr.id,
                        mgr.phone,
                        mgr.fio,
                        mgr.tg_id,
                        mgr.username,
                        "Активен" if mgr.status else "Неактивен",
                    ]
                )

            ws_sec = wb.create_sheet("СБ")
            securities = await session.execute(select(Security))
            ws_sec.append(["ID", "Телефон", "ФИО", "TG ID", "Username", "Статус"])
            for sec in securities.scalars():
                ws_sec.append(
                    [
                        sec.id,
                        sec.phone,
                        sec.fio,
                        sec.tg_id,
                        sec.username,
                        "Активен" if sec.status else "Неактивен",
                    ]
                )

            ws_perm = wb.create_sheet("Постоянные пропуска")
            stmt = select(PermanentPass, Resident.fio, Resident.plot_number).join(
                Resident, PermanentPass.resident_id == Resident.id
            )
            passes = await session.execute(stmt)
            ws_perm.append(
                [
                    "ID",
                    "Резидент ID",
                    "ФИО резидента",
                    "Участок",
                    "Марка",
                    "Модель",
                    "Номер",
                    "Владелец",
                    "Статус",
                ]
            )
            for pass_data in passes:
                pp = pass_data[0]
                ws_perm.append(
                    [
                        pp.id,
                        pp.resident_id,
                        pass_data[1],
                        pass_data[2],
                        pp.car_brand,
                        pp.car_model,
                        pp.car_number,
                        pp.car_owner,
                        pp.status,
                    ]
                )

            ws_temp = wb.create_sheet("Временные пропуска")
            headers = [
                "ID",
                "Тип владельца",
                "ФИО",
                "Участок/Компания",
                "Должность",
                "Тип ТС",
                "Категория веса",
                "Категория длины",
                "Номер авто",
                "Марка",
                "Груз",
                "Цель",
                "Дата визита",
                "Статус",
            ]
            ws_temp.append(headers)

            res_stmt = (
                select(TemporaryPass, Resident.fio, Resident.plot_number)
                .join(Resident, TemporaryPass.resident_id == Resident.id)
                .where(TemporaryPass.owner_type == "resident")
            )
            res_temp_passes = await session.execute(res_stmt)
            for tp_data in res_temp_passes:
                tp = tp_data[0]
                ws_temp.append(
                    [
                        tp.id,
                        "Резидент",
                        tp_data[1],
                        tp_data[2],
                        "",
                        tp.vehicle_type,
                        tp.weight_category,
                        tp.length_category,
                        tp.car_number,
                        tp.car_brand,
                        tp.cargo_type,
                        tp.purpose,
                        tp.visit_date.strftime("%Y-%m-%d"),
                        tp.status,
                    ]
                )

            contr_stmt = (
                select(
                    TemporaryPass,
                    Contractor.fio,
                    Contractor.company,
                    Contractor.position,
                )
                .join(Contractor, TemporaryPass.contractor_id == Contractor.id)
                .where(TemporaryPass.owner_type == "contractor")
            )
            contr_temp_passes = await session.execute(contr_stmt)
            for tp_data in contr_temp_passes:
                tp = tp_data[0]
                ws_temp.append(
                    [
                        tp.id,
                        "Подрядчик",
                        tp_data[1],
                        tp_data[2],
                        tp_data[3],
                        tp.vehicle_type,
                        tp.weight_category,
                        tp.length_category,
                        tp.car_number,
                        tp.car_brand,
                        tp.cargo_type,
                        tp.purpose,
                        tp.visit_date.strftime("%Y-%m-%d"),
                        tp.status,
                    ]
                )

            buffer = BytesIO()
            wb.save(buffer)
            buffer.seek(0)
            uid = event.callback.user.user_id
            await bot.send_message(
                user_id=uid,
                text="📊 Экспорт статистики завершен",
                attachments=[
                    InputMediaBuffer(
                        buffer.getvalue(),
                        filename="Статистика.xlsx",
                        type=UploadType.FILE,
                    )
                ],
            )
            await show_statistics(event)
    except Exception as e:
        await bot.send_message(user_id=RAZRAB, text=f"{event.callback.user.user_id} - {e!s}")
        await asyncio.sleep(0.05)

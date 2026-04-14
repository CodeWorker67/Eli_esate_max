import asyncio
import os
from sqlalchemy.ext.asyncio import create_async_engine, AsyncSession
from sqlalchemy.orm import sessionmaker
from sqlalchemy import insert
from openpyxl import load_workbook
from db import models


async def import_tables_from_excel():
    # Создаем директорию для экспорта (если не существует)
    os.makedirs("export_import", exist_ok=True)

    # Создаем асинхронное подключение к БД
    engine = create_async_engine(models.con_string)
    AsyncSessionLocal = sessionmaker(
        engine, class_=AsyncSession, expire_on_commit=False
    )

    async with AsyncSessionLocal() as session:
        # Получаем список всех таблиц
        tables = models.Base.metadata.tables

        for table_name, table_obj in tables.items():
            excel_file = f"export_import/{table_name}.xlsx"

            # Проверяем существование файла
            if not os.path.exists(excel_file):
                print(f"Файл {excel_file} не найден. Пропускаем.")
                continue

            # Загружаем книгу Excel
            wb = load_workbook(excel_file, read_only=True, data_only=True)
            sheet = wb.active

            # Получаем заголовки таблицы
            headers = []
            for cell in sheet[1]:
                headers.append(cell.value)

            # Получаем названия колонок в таблице БД
            table_columns = [column.name for column in table_obj.columns]

            # Фильтруем заголовки, оставляя только существующие в таблице
            valid_headers = []
            for header in headers:
                if header in table_columns:
                    valid_headers.append(header)

            # Если нет валидных колонок, пропускаем файл
            if not valid_headers:
                print(f"В файле {excel_file} нет подходящих колонок. Пропускаем.")
                wb.close()
                continue

            # Собираем данные для вставки
            data = []
            for row in sheet.iter_rows(min_row=2, values_only=True):
                # Пропускаем пустые строки
                if all(cell is None for cell in row):
                    continue

                row_dict = {}
                valid_row = False

                for i, header in enumerate(headers):
                    # Обрабатываем только валидные заголовки
                    if header not in valid_headers:
                        continue

                    value = row[i] if i < len(row) else None

                    # Если хотя бы одно значение не None - строка валидна
                    if value is not None:
                        valid_row = True

                    row_dict[header] = value

                if valid_row:
                    data.append(row_dict)

            wb.close()

            if data:
                # Выполняем массовую вставку
                await session.execute(insert(table_obj), data)
                await session.commit()
                print(f"Импортировано {len(data)} строк в таблицу {table_name}")
            else:
                print(f"Файл {excel_file} не содержит данных. Пропускаем.")

    await engine.dispose()


if __name__ == "__main__":
    asyncio.run(import_tables_from_excel())